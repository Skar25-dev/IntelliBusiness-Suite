import pandas as pd
from datetime import datetime
from pathlib import Path
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side # Añadimos Border y Side

# Localización de la carpeta de reportes en la raíz
BASE_DIR = Path(__file__).resolve().parent.parent.parent
REPORTS_DIR = BASE_DIR / "reports"

def generate_excel_report(df: pd.DataFrame, report_name: str, sheet_name: str = "Reporte") -> str:
    REPORTS_DIR.mkdir(exist_ok=True)

    df.columns = [str(col).replace('_', ' ').upper() for col in df.columns]

    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    filename = f"{report_name}_{timestamp}.xlsx"
    filepath = REPORTS_DIR / filename

    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]

        # --- 1. DEFINIR EL ESTILO DE LA CUADRÍCULA ---
        thin_side = Side(border_style="thin", color="CCCCCC")
        cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        # --- 2. ESTILOS DE CABECERA ---
        header_fill = PatternFill(start_color="306bac", end_color="306bac", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        # --- 3. APLICAR ESTILOS A TODAS LAS CELDAS ---
        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, 
                                      min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.border = cell_border
                if cell.row == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

        # --- 4. AJUSTE DE COLUMNAS Y FORMATO DE MONEDA ---
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            header_name = str(column[0].value).lower()

            for cell in column:
                if any(word in header_name for word in ['amount', 'precio', 'total', 'cost', 'monto']):
                    cell.number_format = '#,##0.00€'

                try:
                    val_len = len(str(cell.value))
                    if val_len > max_length:
                        max_length = val_len
                except: pass
            
            adjusted_width = min(max(max_length + 2, 10), 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    return str(filepath)